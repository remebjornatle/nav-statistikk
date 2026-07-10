#!/usr/bin/env python3
"""
Scrapes NAV statistics pages, downloads and parses Excel files, writes data.json.
Run locally or via GitHub Actions (workflow_dispatch).
"""
import json
import re
import urllib.parse
from datetime import date
from io import BytesIO

import requests
import openpyxl

BASE = "https://www.nav.no"
HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; nav-statistikk-dashboard/1.0)"}

UF_PAGE = (
    f"{BASE}/no/nav-og-samfunn/statistikk/"
    "aap-nedsatt-arbeidsevne-og-uforetrygd-statistikk/uforetrygd/diagnoser-uforetrygd"
)
SF_PAGE = (
    f"{BASE}/no/nav-og-samfunn/statistikk/"
    "sykefravar-statistikk/sykefravaersstatistikk-arsstatistikk"
)
AAP_PAGE = (
    f"{BASE}/no/nav-og-samfunn/statistikk/"
    "aap-nedsatt-arbeidsevne-og-uforetrygd-statistikk/arbeidsavklaringspenger"
)


# ── Helpers ──────────────────────────────────────────────────────────────────

def find_xlsx(page_url, keyword):
    """Return the full URL of the first xlsx link on page whose filename contains keyword."""
    print(f"  Scraping {page_url} for '{keyword}'...")
    r = requests.get(page_url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    links = re.findall(r'/_/attachment/download/[^"\'>\s]+\.xlsx', r.text)
    seen = set()
    for link in links:
        decoded = urllib.parse.unquote(link)
        if keyword.lower() in decoded.lower() and link not in seen:
            seen.add(link)
            print(f"    Found: ...{decoded[-60:]}")
            return BASE + link
    raise ValueError(f"No xlsx matching '{keyword}' on {page_url}")


def download_wb(url):
    print(f"  Downloading ...{url[-60:]}")
    r = requests.get(url, headers=HEADERS, timeout=60)
    r.raise_for_status()
    return openpyxl.load_workbook(BytesIO(r.content), data_only=True)


def num(v):
    """Return a numeric value (int or rounded float) or None."""
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return round(v, 2)
    return None


def find_year_col(header_row):
    """Return (col_index, [years]) for the first year found in a row."""
    for i, v in enumerate(header_row):
        if isinstance(v, int) and 2000 < v < 2030:
            years = [w for w in header_row[i:] if isinstance(w, int) and 2000 < w < 2030]
            return i, years
    raise ValueError("No year column found in header row")


# ── Uføretrygd ───────────────────────────────────────────────────────────────

def parse_uforetrygd(wb1, wb3):
    # ── Tabell 1: diagnosis groups 2012-2025 ────────────────────────────────
    ws1 = wb1["Antall"]
    rows1 = list(ws1.iter_rows(values_only=True))

    # Find year header (first row containing 2012)
    yr_row = next(i for i, r in enumerate(rows1) if 2012 in r)
    year_col, years1 = find_year_col(rows1[yr_row])
    n1 = len(years1)

    # Collect "Kvinner og menn" section label→data (stop at "Menn" section)
    section = {}
    for row in rows1[yr_row + 1:]:
        if row[1] == "Menn":
            break
        label = row[0]
        if not isinstance(label, str) or not label.strip():
            continue
        label = label.strip()
        vals = [num(row[year_col + i]) for i in range(n1)]
        section[label] = vals

    # Map friendly name → search fragment (case-insensitive contains)
    GROUP_MAP = [
        ("Nervesystemet (inkl. ME)",    "sykdommer i nervesystemet"),
        ("Psykiske lidelser (total)",   "psykiske lidelser og atferdsforstyrrelser"),
        ("Nevroser/atferdsforstyrrelser", "nevroser og atferdsforstyrrelser"),
        ("Personlighetsforstyrrelser",  "personlighetsforstyrrelser"),
        ("Psykisk utviklingshemming",   "psykisk utviklingshemming"),
        ("Medfødte misdannelser",       "medfødte misdannelser"),
        ("Myalgi/fibromyalgi",          "myalgi/fibromyalgi"),
        ("Muskel-skjelett (total)",     "muskel-skjelettsystemet og bindevev"),
        ("Skader/forgiftning",          "skader, forgiftninger"),
    ]

    def find_label(fragment):
        for key, vals in section.items():
            if fragment.lower() in key.lower():
                return vals
        return None

    groups1 = [
        {"label": name, "data": find_label(frag)}
        for name, frag in GROUP_MAP
        if find_label(frag) is not None
    ]

    # ── Tabell 3: individual diagnoses 2017-2025 (absolute numbers) ─────────
    ws3 = wb3["Antall"]
    rows3 = list(ws3.iter_rows(values_only=True))

    # Two year-header rows: first is %, second is absolute counts
    yr_rows3 = [i for i, r in enumerate(rows3) if r[2] and isinstance(r[2], int) and 2000 < r[2] < 2030]
    abs_yr_idx = yr_rows3[1] if len(yr_rows3) > 1 else yr_rows3[0]
    year_col3, years3 = find_year_col(rows3[abs_yr_idx])
    n3 = len(years3)

    diag3 = []
    for row in rows3[abs_yr_idx + 1:abs_yr_idx + 15]:
        code, label = row[0], row[1]
        if not isinstance(label, str) or not label.strip():
            continue
        vals = [num(row[year_col3 + i]) for i in range(n3)]
        if any(v is not None for v in vals):
            diag3.append({
                "label": label.strip(),
                "code": str(code).strip() if code else "",
                "data": vals,
            })

    # ── Tabell 2 (Tilgang): new ME recipients ────────────────────────────────
    # Tilgang sheet also has two blocks: % first, absolute second — use second
    wt = wb3["Tilgang"]
    rows_t = list(wt.iter_rows(values_only=True))
    yr_rows_t = [i for i, r in enumerate(rows_t) if r[2] and isinstance(r[2], int) and 2000 < r[2] < 2030]
    abs_yr_t = yr_rows_t[1] if len(yr_rows_t) > 1 else yr_rows_t[0]
    year_col_t, years_t = find_year_col(rows_t[abs_yr_t])
    nt = len(years_t)

    new_me = []
    for row in rows_t[abs_yr_t + 1:abs_yr_t + 15]:
        label = row[1]
        if isinstance(label, str) and "utmattelse" in label.lower():
            new_me = [num(row[year_col_t + i]) for i in range(nt)]
            break

    # ── Total I ALT (for KPI cards) ─────────────────────────────────────────
    total_series = find_label("I ALT") or find_label("i alt")

    return {
        "years1": years1,
        "years3": years3,
        "total": total_series,
        "diag3": diag3,
        "newME": new_me,
        "groups1": groups1,
    }


# ── Sykefravær ───────────────────────────────────────────────────────────────

def parse_sheet_sf(ws, stop_label="Kvinner"):
    """Parse one sykefravær sheet, stopping at the gender breakdown sections."""
    rows = list(ws.iter_rows(values_only=True))
    yr_row = next(i for i, r in enumerate(rows) if any(isinstance(v, int) and 2000 < v < 2030 for v in r))
    year_col, years = find_year_col(rows[yr_row])
    n = len(years)

    result = []
    for row in rows[yr_row + 1:]:
        if row[0] == stop_label:
            break
        group, code, label = row[0], row[1], row[2]
        if not isinstance(label, str) or not label.strip():
            continue
        vals = [num(row[year_col + i]) for i in range(n)]
        if any(v is not None for v in vals):
            result.append({
                "group": str(group).strip() if group else "",
                "code":  str(code).strip()  if code  else "",
                "label": label.strip(),
                "data":  vals,
            })
    return years, result


def parse_sykefravar(wb):
    EXCLUDE_LABELS = {"Alle tapte dagsverk"}
    EXCLUDE_CODES  = {"R992", ""}          # Skip COVID and total row
    PSYK_CODES     = {"P02", "P03", "P29", "P74", "P76"}

    years_rate, rows_rate = parse_sheet_sf(wb["Andel per 10 000 avtalte dagsve"])
    years_abs,  rows_abs  = parse_sheet_sf(wb["Tapte dagsverk"])

    def keep(r):
        return r["code"] not in EXCLUDE_CODES and r["label"] not in EXCLUDE_LABELS

    return {
        "years":     years_rate,
        "diag_rate": [r for r in rows_rate if keep(r)],
        "diag_abs":  [r for r in rows_abs  if keep(r)],
        "psyk":      [r for r in rows_rate if r["code"] in PSYK_CODES],
    }


# ── AAP ──────────────────────────────────────────────────────────────────────

def parse_aap(wb):
    KNOWN_GROUPS = {
        'Allment', 'Andre lidelser', 'Hjerte- og karsykdommer',
        'Muskel- og skjelettlidelser', 'Psykiske lidelser',
        'Sykdom i fordøyelsesorganene', 'Sykdommer i luftveiene',
        'Sykdommer i nervesystemet', 'Kreft',
    }

    # ── Group totals from Hovedgruppe sheet ──────────────────────────────────
    ws_h = wb['1.Hovedgruppe']
    rows_h = list(ws_h.iter_rows(values_only=True))
    hdr_h = next(r for r in rows_h if r[2] == 'Jan')
    mai_col_h = next(i for i, v in enumerate(hdr_h) if v == 'Mai')
    total = None
    groups = []
    for row in rows_h:
        if not isinstance(row[1], str):
            continue
        label = row[1].strip()
        val = row[mai_col_h]
        if not isinstance(val, (int, float)):
            continue
        if label == 'I alt':
            total = int(val)
        elif label in KNOWN_GROUPS:
            groups.append({'label': label, 'count': int(val)})

    # ── Individual diagnoses from detail sheet ───────────────────────────────
    ws = wb['1a. Diagnose kode navn antall']
    rows = list(ws.iter_rows(values_only=True))
    hdr = next(r for r in rows if r[3] == 'Jan')
    mai_col = next(i for i, v in enumerate(hdr) if v == 'Mai')

    period_row = next(r for r in rows if isinstance(r[1], str) and r[1].startswith('Periode'))
    period = period_row[1].strip()

    current_group = None
    diags = []
    for row in rows:
        v = row[1]
        if not isinstance(v, str):
            continue
        v = v.strip()
        if v in KNOWN_GROUPS:
            current_group = v
            continue
        if not current_group:
            continue
        parts = v.split(' ', 1)
        if (len(parts) == 2 and len(parts[0]) >= 3
                and parts[0][0].isalpha() and parts[0][1].isdigit()):
            mai_val = row[mai_col]
            if isinstance(mai_val, (int, float)):
                diags.append({
                    'code':  parts[0],
                    'label': parts[1],
                    'group': current_group,
                    'count': int(mai_val),
                })

    diags.sort(key=lambda x: x['count'], reverse=True)
    return {'period': period, 'total': total, 'groups': groups, 'diags': diags}


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("=== Fetching Excel URLs ===")
    t1_url  = find_xlsx(UF_PAGE,  "Upantall_diagnoser")
    t3_url  = find_xlsx(UF_PAGE,  "mest brukte")
    sf_url  = find_xlsx(SF_PAGE,  "SYFRA 2301")
    aap_url = find_xlsx(AAP_PAGE, "AAP180")

    print("\n=== Downloading & parsing uføretrygd ===")
    wb1 = download_wb(t1_url)
    wb3 = download_wb(t3_url)
    uforetrygd = parse_uforetrygd(wb1, wb3)

    print("\n=== Downloading & parsing sykefravær ===")
    wb_sf = download_wb(sf_url)
    sykefravar = parse_sykefravar(wb_sf)

    print("\n=== Downloading & parsing AAP ===")
    wb_aap = download_wb(aap_url)
    aap = parse_aap(wb_aap)

    out = {
        "generated": date.today().isoformat(),
        "sources": {
            "uforetrygd_page": UF_PAGE,
            "sykefravar_page": SF_PAGE,
            "aap_page":        AAP_PAGE,
            "tabell1_url":  t1_url,
            "tabell3_url":  t3_url,
            "syfra2301_url": sf_url,
            "aap180_url":   aap_url,
        },
        "uforetrygd": uforetrygd,
        "sykefravar": sykefravar,
        "aap": aap,
    }

    with open("data.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    # Bake data inline into index.html so the page loads instantly (no fetch needed)
    inline = json.dumps(out, ensure_ascii=False, separators=(',', ':'))
    marker_begin = "// @@NAV_DATA_BEGIN@@"
    marker_end   = "// @@NAV_DATA_END@@"
    html_path = "index.html"
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    i0 = html.index(marker_begin)
    i1 = html.index(marker_end) + len(marker_end)
    new_block = f"{marker_begin}\nconst NAV_DATA = {inline};\n{marker_end}"
    html = html[:i0] + new_block + html[i1:]
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  index.html updated with inline data")

    print(f"\n=== Done ===")
    print(f"  data.json written")
    print(f"  Uføretrygd years:  {uforetrygd['years3']}")
    print(f"  Sykefravær years:  {sykefravar['years']}")
    me = next((d for d in uforetrygd["diag3"] if "utmattelse" in d["label"].lower()), None)
    print(f"  ME data:           {me['data'] if me else 'NOT FOUND'}")


if __name__ == "__main__":
    main()
